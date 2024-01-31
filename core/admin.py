from unicodedata import category
from django.contrib import admin
import os 
import requests
import pandas as pd
import openpyxl
import csv
from .models import *
import time
import re
import ast
import json
import random
from openpyxl import load_workbook
# Register your models here.
from PIL import Image
# from io import BytesIO
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils.text import slugify
from import_export.admin import ExportActionMixin, ImportExportMixin, ImportMixin
# import io
# from import_export.resources import ModelResource
from django.http import HttpResponse
# import openpyxl
# from openpyxl.writer.excel import save_virtual_workbook

def make_refund_accepted(modeladmin, request, queryset):
    queryset.update(refund_requested=False, refund_granted=True)


make_refund_accepted.short_description = 'Update orders to refund granted'



class AddressAdmin(admin.ModelAdmin):
    list_display = [
        'user',
        'street_address',
        'apartment_address',
        'country',
        'zip',
        'address_type',
        'default'
    ]
    list_filter = ['default', 'address_type', 'country']
    search_fields = ['user', 'street_address', 'apartment_address', 'zip']


def copy_items(modeladmin, request, queryset):
    for object in queryset:
        object.id = None
        object.save()


copy_items.short_description = 'Copy Items'



@admin.register(ExcelFile)
class ExcelFileAdmin(admin.ModelAdmin):
    list_display = ('name', 'file')

    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_images')
    os.makedirs(temp_dir, exist_ok=True)

     


    def save_model(self, request, obj, form, change):
        obj_file_name= obj.file
        gender_type = obj.gender
        df = pd.read_excel(obj_file_name,  engine='openpyxl')

        
        if obj.label == "A":

            for index, row in df.iterrows():
                try:
                    id = int(row["id"])
                except:
                    pass
                name =row["nom d'article"]
                # images = sheet["E" + str(index)]
                brand_name = row["brand_name"]
                brand_id = row['brand_id']
                # product_Group_Id = sheet["F" + str(index)]
                try:
                    sellingPrice = int(row['prix'])
                    discountPrice = int(row['ancien prix'])
                except:
                    sellingPrice=0
                    discountPrice=0

                try:
                    image_src = row["images for website"]
                    # image_src = row['images_prod']
                    
                except:
                    image_src = row['chemins photos']

                if not pd.isna(row['code produit']):
                    article_id = row['code produit']
                else:
                    try:
                        article_id = name.split(" ")[-1]
                    except:
                        article_id=str(name).split(" ")[-1]
                    
                
                sizes = row['sizes_exist']
                if sizes:
                    sizes = sizes
                else:
                    sizes="not"
                
                sizes_not_exist = row['sizes_not_exist']
                if sizes_not_exist:
                    sizes_not_exist =  sizes_not_exist
                else:
                    sizes_not_exist="not"
                    


                categoryName = row["category_name"]
                
                description = row["details d'article"]
                try:
                    idxn=int(row["id"])
                except:
                    pass
                try:
                        
                    shipping = row["delivery"]
                except:
                    shipping = row["duree de livraison"]
                rating = 0
                # if rating:
                #     rating = float(rating)
                # else:
                #     rating = 0
                # stock = row["stock"]
                stock = +99
                # Use regular expression to extract only the numbers
                # if stock:
                #     pass
                    
                # else:
                #     stock = 0

                #in past it was row['product_details_attr'] now it is like bellow
                details_upp=row['product_details_attr'].replace('d"', "de ")
                details_upp=details_upp.replace("/","-")
                details_upp=details_upp.replace('%','percent')
                details_upp=details_upp.replace("d’","de ")
                details_upp=details_upp.replace("d'","de ")
                details_upp=details_upp.replace("'",'"')
                details_upp=details_upp.replace(" ,",'",')
                details = details_upp




                try:
                    image_src = row["images for website"]
                    # image_src = row['images_prod']
                except:
                    image_src = row['chemins photos']

                
                if image_src :
                    # image_src ="https://cdn.dsmcdn.com"+ image_src.split(",")[0].replace("\\","/").replace(" ","")
                    # image_folder = image_src.split("/")[-2]
                    # image_folder = os.path.join(str(idxn), image_folder)
                    try:
                        image_src=image_src.split(",")[0].replace("\\","/").replace(" ","")
                    except:
                        image_src=str(image_src).split(",")[0].replace("\\","/").replace(" ","")
                    local_image_path = image_src 
                    # Replace this with the actual path of the image on your computer

                    # Construct the path to the media_root directory
                    media_root = os.path.join(settings.MEDIA_ROOT, 'images')
                    # Construct the path to the media_root directory
                    media_root1 = os.path.join(settings.MEDIA_ROOT,"images_upload_product" )

                    # Create the 'images' directory if it doesn't exist
                    if not os.path.exists(media_root):
                        os.makedirs(media_root)
                    
                    if not os.path.exists(media_root1):
                        os.makedirs(media_root1)

                    # Get the image filename from the local path
                    filename =  local_image_path 
                    image_path = os.path.join(media_root,media_root1,  filename.replace("+","") )
                
                    
                    # Copy the image file to the media_root directory
                    # response=requests.get(local_image_path)
                    
                    # if response.status_code == 200 :
                        
                    #     # with open(image_path, 'wb') as dest_file:
                            
                    #     #     dest_file.write(response.content)
                    #     image_data = response.content
                    #     image = Image.open(io.BytesIO(image_data))
                    #     rgba_image = image.convert("RGBA")

                    #     # Create a file path for saving the WebP image
                    #     webp_file_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")

                    #     # Save the RGBA image as WebP
                    #     rgba_image.save(webp_file_path, "WEBP")
                            
                        

                

                            
                    
                    if categoryName:
                        
                        doesexist = Category.objects.filter(title=categoryName)
                        if len(doesexist) > 0:
                            categoryName=doesexist.first().title
                        else:
                            Category.objects.create(
                                title=categoryName,
                                slug=slugify(categoryName),
                                
                                image=image_path.replace("/root/demo/media_root","") 
                            )


                    if image_src:
                        if name:

                            slug = slugify(name)
                            queryset = Item.objects.filter(slug=slug)
                            numbr = 1
                            new_slug = slug
                            random_number = random.randint(0, 10000)
                            if queryset.filter(slug=new_slug).exists():
                            
                                new_slug = f"{slug}-{ idxn }"
                                if queryset.filter(slug=new_slug).exists():
                                    new_slug = f"{slug}-{ random_number }"


                                
                           
                            
                            

                                
                            if discountPrice > sellingPrice:
                                label_pro="P"
                            else:
                                label_pro="N"


                            doesexist11 = Item.objects.filter(id_item = idxn)
                            imagepathitem = os.path.join(media_root,"images_upload_product",media_root1,  filename.replace("/root/demo/media_root",""  ))
                            if len(doesexist11) == 0:
                                itemnow = Item.objects.create(
                                    id_item=idxn,
                                    title=name,
                                    slug=  new_slug,
                                    price=sellingPrice,
                                    discount_price=discountPrice,
                                    brand_name=brand_name,
                                    category=Category.objects.filter(title=categoryName).first(),
                                    label=label_pro,
                                    article_id=article_id,
                                    stock_no=stock,
                                    details= details ,
                                    gender=gender_type,
                                    description_short=description,
                                    description_long=description,
                                    tags= details ,
                                   
                                    rating=rating,
                                    color_exist="M,S",
                                    color_not_exist="F",
                                    size_exist=sizes,
                                    size_not_exist=sizes_not_exist,
                                    wishlist_num=0,
                                    opinion_num=0,
                                    shipping=shipping,


                                )
                                
                                
                                
                                try:
                                    # image_src = row['images_prod']

                                    image_src = row["images for website"]
                                     
                                    for image_url in  image_src.split(",")[1:]:
                                        # image_src = "https://cdn.dsmcdn.com"+ image_url.replace(" ","")
                                        # image_folder = image_src.split("/")[-3]
                                        
                                        local_image_path = image_url 
                                        # Replace this with the actual path of the image on your computer

                                        
                                        
                                        # Create the 'images' directory if it doesn't exist
                                        

                                        # Get the image filename from the local path
                                        # filename = os.path.basename(local_image_path)
                                        filename = local_image_path
                                        # print(f"{local_image_path}/{filename}")
                                        image_path = os.path.join(media_root,media_root1,  filename.replace("+","") )
                                        
                                        
                                        # image_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")
                                        slug_ex=itemnow.slug
                                            
                                        
                                        # Copy the image file to the media_root directory
                                        
                                        # ImageItem.objects.create(
                                        #     item=Item.objects.filter(slug=slug_ex).first(),
                                        #     slug=slug_ex,
                                        #     image=image_path.replace("/root/demo/media_root","") 
                                        # )
                                        
                                        # response = requests.get(local_image_path)
                                        # if response.status_code == 200:
                                            
                                        #     # with open(image_path, 'wb') as dest_file:
                                            
                                                
                                        #     #     dest_file.write(response.content)
                                                
                                        #     image_data = response.content
                                        #     image = Image.open(io.BytesIO(image_data))
                                        #     rgba_image = image.convert("RGBA")

                                        #     # Create a file path for saving the WebP image
                                        #     webp_file_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")

                                        #     # Save the RGBA image as WebP
                                        #     rgba_image.save(webp_file_path, "WEBP")
                                except:
                                     
                                    image_src = str(row['chemins photos'])
                                     
                                        
                                    # for image_url in  image_src.split(",")[1:]:
                                    for image_url in  image_src.split(","):
                                        image_src = "https://cdn.dsmcdn.com"+ image_url.replace(" ","")
                                        image_folder = image_src.split("/")[-3]
                                         
                                        image_folder = os.path.join(image_folder, image_src.split("/")[-2])
                                        
                                        local_image_path = image_src 
                                        # Replace this with the actual path of the image on your computer

                                        
                                        
                                        # Create the 'images' directory if it doesn't exist
                                        

                                        # Get the image filename from the local path
                                        filename = os.path.basename(local_image_path)
                                         
                                        image_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")
                                        
                                        
                                        # image_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")
                                        slug_ex=itemnow.slug
                                            
                                        
                                        # Copy the image file to the media_root directory
                                        
                                        # ImageItem.objects.create(
                                        #     item=Item.objects.filter(slug=slug_ex).first(),
                                        #     slug=slug_ex,
                                        #     image=image_path.replace("/root/demo/media_root","") 
                                        # )
                                                
        else:
            for index, row in df.iterrows():
                try:
                    sellingPrice = int(row['prix'])
                except:
                    sellingPrice=0
                try:
                    discountPrice = int(row['ancien prix'])
                except:
                    discountPrice=0
                try:

                    id = int(row["id"])
                except:
                    pass
                try:
                    
                    tar_pro= Item.objects.get(id_item=id)
                     

                except:
                    titleprod=row["nom d'article"]
                    categoryName1=row["category_name"]

                    try:
                        category_instance = Category.objects.get(title=categoryName1)
                    except Category.DoesNotExist:
                        # Handle the case where the category with the specified title doesn't exist
                        category_instance = 0
                    
                    if not pd.isna(row['code produit']):
                        article_id = row['code produit']
                    else:
                        article_id = name.split(" ")[-1]

                    description = row["details d'article"]
                    rating = 0

                    sizes = row['sizes_exist']
                    if sizes:
                        sizes = sizes
                    else:
                        sizes="not"

                    sizes_not_exist = row['sizes_not_exist']
                    if sizes_not_exist:
                        sizes_not_exist =  sizes_not_exist
                    else:
                        sizes_not_exist="not"


                    if discountPrice > sellingPrice:
                        label_pro="P"
                    else:
                       label_pro="N"
                    
                    tar_pro= Item.objects.create(id_item=id,title=titleprod,category=category_instance, article_id=article_id,details= details ,label=label_pro
                                                 ,description_short=description,description_long=description,rating=rating,tags= details ,
                                                 size_exist=sizes,size_not_exist=sizes_not_exist,price=sellingPrice,discount_price=discountPrice,)
                    price1=row["prix"]
                    tar_pro.price=int(price1)
                    try:
                        oldprice1=row["ancien prix"]
                    except:
                        oldprice1=row["prix reduction"] 
                    tar_pro.discount_price=oldprice1
                     
                    
                    
                    brand_name = row["brand_name"]
                   
                    tar_pro.brand_name=brand_name

                    slug = slugify(titleprod)
                    queryset = Item.objects.filter(slug=slug)
                    numbr = 1
                    new_slug = slug
                    random_number = random.randint(0, 10000)
                    if queryset.filter(slug=new_slug).exists():
                            
                        new_slug = f"{slug}-{ id }"
                        if queryset.filter(slug=new_slug).exists():
                            new_slug = f"{slug}-{ random_number }"


                    tar_pro.slug=new_slug
                    

   
                                     
                if discountPrice > sellingPrice:
                        label_pro="P"
                else:
                    label_pro="N"
                tar_pro.label=label_pro
                try:
                    if not pd.isna(row["nom d'article arabe"])  :
                        arabe_title = row["nom d'article arabe"]
                        tar_pro.title_ar=arabe_title
                except:
                    pass
                
                
                try:
                    if not pd.isna(row["nom d'article anglais"])    :
                        english_title = row["nom d'article anglais"]
                        tar_pro.title_en=english_title
                except:
                    pass
                
                try:
                    if not pd.isna(row["product_details_attr_ar"])  :
                        det_ar = row["product_details_attr_ar"]
                        tar_pro.details_ar=det_ar
                except:
                    pass
                
                try:
                    if not pd.isna(row["product_details_attr_en"])  :
                        det_en = row["product_details_attr_en"]
                        tar_pro.details_en=det_en
                except:
                    pass
                
                try:
                    if not pd.isna(row["details d'article"])  :
                        desc = row["details d'article"]
                        tar_pro.description_long=desc
                except:
                    pass
                try:
                    if not pd.isna(row["details d'article arabe"])  :
                        desc_ar = row["details d'article arabe"]
                        tar_pro.description_long_ar=desc_ar
                     
                except:
                    pass
                
                try:
                    if not pd.isna(row["details d'article anglais"])  :
                        desc_en = row["details d'article anglais"]
                        tar_pro.description_long_en=desc_en
                    
                except:
                    pass
                
                # if row["id"]:
                     
                    # image_src = row['chemins photos']
                    # image_src = row["images for website"]
                
                    # if image_src :
                    #     # image_src ="https://cdn.dsmcdn.com"+ image_src.split(",")[0].replace("\\","/").replace(" ","")
                    #     # image_folder = image_src.split("/")[-2]
                    #     # image_folder = os.path.join(str(row["id"]), image_folder)
                    #     local_image_path = image_src.split(",")[0].replace("\\","/").replace(" ","") 
                    #     # Replace this with the actual path of the image on your computer

                    #     # Construct the path to the media_root directory
                    #     media_root = os.path.join(settings.MEDIA_ROOT, 'images')
                    #     # Construct the path to the media_root directory
                    #     media_root1 = os.path.join(settings.MEDIA_ROOT,"images_upload_product" )


                    #     # image_src ="https://cdn.dsmcdn.com"+ image_src.split(",")[0].replace("\\","/").replace(" ","")
                    #     # image_folder = image_src.split("/")[-2]
                    #     # image_folder = os.path.join(image_src.split("/")[-3], image_folder)
                    #     # local_image_path = image_src 
                    #     # Replace this with the actual path of the image on your computer

                    #     # Construct the path to the media_root directory
                    #     # media_root = os.path.join(settings.MEDIA_ROOT, 'images')
                    #     # # Construct the path to the media_root directory
                    #     # media_root1 = os.path.join(settings.MEDIA_ROOT,"images_upload_product", image_folder)

                    #     # Create the 'images' directory if it doesn't exist
                    #     if not os.path.exists(media_root):
                    #         os.makedirs(media_root)
                        
                    #     if not os.path.exists(media_root1):
                    #         os.makedirs(media_root1)

                    #     # Get the image filename from the local path
                    #     # filename = os.path.basename(local_image_path)
                    #     filename =local_image_path
                    #     # image_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")
                    #     image_path = os.path.join(media_root,media_root1,  filename.replace("+","") )

                    #     # imagepathitem = os.path.join(media_root,"images_upload_product",media_root1, f"{filename.split('.')[0]}.webp")
                    #     imagepathitem = os.path.join(media_root,media_root1, filename.replace("+",""))
                    #     tar_pro.image=imagepathitem.replace("/root/demo/media_root","")
                # else:
                #     pass
                try:
                    
                    sizes = row['sizes_exist']
                    tar_pro.size_exist=sizes
                except:
                    pass
                try:
                    size_not_exist = row['sizes_not_exist']
                    tar_pro.size_not_exist=size_not_exist
                except:
                    pass
                    
                try:
                    
                
                    price = row['prix']
                    tar_pro.price=price
                except:
                    pass
                    
                
                try:
                    
                
                    pricerid = row['ancien prix']
                    tar_pro.discount_price=pricerid
                except:
                    pass
                try:
                    try:
                        
                        shipping = row["delivery"]
                    except:
                        shipping = row["duree de livraison"]
                        
                    tar_pro.shipping=shipping
                
                except:
                    pass
                
               
                
                try:
                    details_upp=row['product_details_attr'].replace('d"', "de ")
                    details_upp=details_upp.replace("/","-")
                    details_upp=details_upp.replace('%','percent')
                    details_upp=details_upp.replace("d’","de ")
                    details_upp=details_upp.replace("d'","de ")
                    details_upp=details_upp.replace("'",'"')
                    details_upp=details_upp.replace(" ,",'",')

                    details =details_upp
                    if details != "NaN":
                        tar_pro.details=details
                    else:
                        tar_pro.details=""


                    
                
                except:
                    pass


                try:
                    if ImageItem.objects.filter(item__id_item=id):
                        # ImageItem.objects.filter(item__id_item=id).delete()
                        # image_src1 = row["chemins photos"]
                        image_src1=row["images for website"]

                        for image_url in  image_src1.split(","):
                            image_src = "https://cdn.dsmcdn.com" + image_url.replace(" ","")
                            image_folder = image_src.split("/")[-3]
                                            
                        
                                            
                            local_image_path = image_src 
                                        
                            # filename = os.path.basename(local_image_path)
                            filename=image_url.replace("+","")
                            media_root = os.path.join(settings.MEDIA_ROOT, 'images')
                            # Construct the path to the media_root directory
                            media_root1 = os.path.join(settings.MEDIA_ROOT,"images_upload_product" )
                                            
                            # image_path = os.path.join(media_root,media_root1, f"{filename.split('.')[0]}.webp")
                            image_path = os.path.join(media_root,media_root1,  filename )
                                            
                                            
                                            
                            slug_ex=Item.objects.filter(id_item=id).first().slug
                                            
                            # ImageItem.objects.create(
                            #     item=Item.objects.filter(id_item=id).first(),
                            #     slug=slug_ex,
                            #     image=image_path.replace("/root/demo/media_root","") 
                            # )
                except:
                    pass           
                
                # else:
                #     # image_src1 = row["chemins photos"]
                #     image_src1=row["images for website"]
                #     for image_url in  image_src1.split(","):
                #         # image_src = "https://cdn.dsmcdn.com"+ image_url.replace(" ","")
                #         # image_folder = image_src.split("/")[-3]
                                         
                #         # image_folder = os.path.join(image_folder, image_src.split("/")[-2])
                                        
                #         local_image_path = image_src 
                                       
                #         # filename = os.path.basename(local_image_path)
                #         filename=image_url.replace("+","")

                #         image_src ="https://cdn.dsmcdn.com"+ image_src.split(",")[0].replace("\\","/").replace(" ","")
                #         image_folder = image_src.split("/")[-3]
                #         image_folder = os.path.join(str(id), image_folder)
                #         local_image_path = image_src 
                #         # Replace this with the actual path of the image on your computer

                #         # Construct the path to the media_root directory
                #         media_root = os.path.join(settings.MEDIA_ROOT, 'images')
                #         # Construct the path to the media_root directory
                #         media_root1 = os.path.join(settings.MEDIA_ROOT,"images_upload_product" )
                                         
                #         image_path = os.path.join(media_root,media_root1,  filename )
                                        
                                        
                                        
                #         slug_ex=Item.objects.filter(id_item=id).first().slug
                                        
                        # ImageItem.objects.create(
                        #     item=Item.objects.filter(id_item=id).first(),
                        #     slug=slug_ex,
                        #     image=image_path.replace("/root/demo/media_root","") 
                        # )
                
                try:
                    tar_pro.save()
                except:
                    pass
                print("modified")
                                     

                                    
                

                    
                    
                    

            
        super().save_model(request, obj, form, change)
             



 

from datetime import timedelta


def add_coupon_to_selected(modeladmin, request, queryset):
    # Assuming you have a Coupon model and a 'coupon_code' field in it
    coupon_code = "MATACOR DZ"  # Replace with your desired coupon code

    # Assuming you have a 'coupon_end_date' field in your Product model
    # and you want to set the coupon to expire in 30 days from today
    coupon_end_date = timezone.now().date() + timedelta(days=30)

    # Apply the coupon to the selected products
    for product in queryset:
        product.coupon_code = coupon_code
        product.coupon_start_date = timezone.now().date()
        product.coupon_end_date = coupon_end_date
        product.has_coupon=True
        product.show_coupon=True
        product.price_after_coupon=80
        product.save()

    modeladmin.message_user(request, "Coupon added to selected items successfully.")

add_coupon_to_selected.short_description = "Add coupon to selected items"

def delete_coupon_to_selected(modeladmin, request, queryset):
    # Assuming you have a Coupon model and a 'coupon_code' field in it
    coupon_code = ""  # Replace with your desired coupon code

    # Assuming you have a 'coupon_end_date' field in your Product model
    # and you want to set the coupon to expire in 30 days from today
     

    # Apply the coupon to the selected products
    for product in queryset:
        product.coupon_code = coupon_code
        product.has_coupon=False
        product.show_coupon=False
        product.save()

    modeladmin.message_user(request, "Coupon deleted to selected items successfully.")

delete_coupon_to_selected.short_description = "delete  coupon to selected items"
 
 
 
def move_to_promotion(modeladmin, request, queryset):
    
    for product in queryset:
        product.label = "P"
         
        product.save()

    modeladmin.message_user(request, "moved to Promotion successfully.")

move_to_promotion.short_description = "Move to Promotion"
 
 
def move_to_new(modeladmin, request, queryset):
    
    for product in queryset:
        product.label = "N"
         
        product.save()

    modeladmin.message_user(request, "moved to New successfully.")

move_to_new.short_description = "Move to New"


def move_to_Sale(modeladmin, request, queryset):
    
    for product in queryset:
        product.label = "S"
         
        product.save()

    modeladmin.message_user(request, "moved to best Sale successfully.")

move_to_Sale.short_description = "Move to best Sale"

class ItemAdmin(admin.ModelAdmin):
    list_display = [
        'title',
        'category','label'
    ]
    list_filter = ['title', 'category','brand_name', 'article_id','id_item','label']
    search_fields = ['title', 'description_long' ,'brand_name' ,'article_id','id_item']
    prepopulated_fields = {"slug": ("title",)}
    actions = [add_coupon_to_selected,delete_coupon_to_selected,move_to_Sale,move_to_new,move_to_promotion]
    list_display = ('id_item','get_image_display','title','price','discount_price','has_coupon', 'coupon_code','label')  # Customize the displayed columns in the admin list view
    

    def get_image_display(self, obj):
        try:
            return mark_safe(f'<img src="{obj.image.url}" width="50" height="50" />')
        except:
            return mark_safe(f'<img src="" width="50" height="50" />')


    get_image_display.allow_tags = True
    get_image_display.short_description = 'Image Preview'
    

    def save_model(self, request, obj, form, change):
        
        sizes_list=[]
        colors_list=[]
        sizes_ex=""
        colors_ex=""
        colors_list_not1=[]
        sizes_list_not1=[]
        if  obj.color_exist:
            colors_list = obj.color_exist.replace(" ","").split(',')
            colors_list_not1 = obj.color_not_exist
            colors_ex= obj.color_exist
        
        if  obj.size_exist:
            sizes_list = obj.size_exist.replace(" ","").split(',')
            sizes_list_not1 = obj.size_not_exist
            sizes_ex= obj.size_exist
        
        if sizes_list_not1 and sizes_list:
            list1 = set(sizes_list_not1.replace(" ","").split(','))
            list2 = set(sizes_list)

            unique_to_list1 = [item for item in list1 if item not in list2]
            unique_to_list2 = [item for item in list2 if item not in list1]


            
            sizes_ex = ",".join(unique_to_list2)

 
        # Add each color to the colors array
        
        
        
        if colors_list_not1 and colors_list:
            list1 = set(colors_list_not1.replace(" ","").split(','))
            list2 = set(colors_list)
           
            unique_to_list1 = [item for item in list1 if item not in list2]
            unique_to_list2 = [item for item in list2 if item not in list1]

 
                            
            colors_ex = ",".join(unique_to_list2)
               

              
        obj.size_exist = sizes_ex
        obj.color_exist = colors_ex
        
                 
        
         
        if obj.has_coupon:
            # If has_coupon is True, set a default value for coupon_code
            obj.coupon_code = obj.coupon_code
        else:
            # If has_coupon is False, set coupon_code to None
            obj.coupon_code = None

        super().save_model(request, obj, form, change)


class CategoryAdmin(admin.ModelAdmin):
    list_display = [
        "get_image_display",
        'title',
        'is_active'
    ]
    list_filter = ['title', 'is_active']
    search_fields = ['title', 'is_active']
    prepopulated_fields = {"slug": ("title",)}

    def get_image_display(self, obj):
            return mark_safe(f'<img src="{obj.image.url}" width="50" height="50" />')

    get_image_display.allow_tags = True
    get_image_display.short_description = 'Image Web'


class CategoryGenTopAdmin(admin.ModelAdmin):
    
    
    prepopulated_fields = {"slug": ("title",)}
    
    def save_model(self, request, obj, form, change):
        if obj.title.lower() in ["hommes", "homme","man","men"]:
            obj.slug = "/shop?gender=M"
        elif obj.title.lower() in ["femmes", "femme","women","woman"]:
            obj.slug = "/shop?gender=F"
        elif obj.title.lower() in ["enfants", "enfant","kids","kid"]:
            obj.slug = "/shop?gender=E"
        else:
            slug1="/shop?topcat="
            try:
                if len(obj.items.all()) == 1:
                  
                    slug1= slug1 +  obj.items.first().slug
                elif len(obj.items.all()) > 1:
                    for item in obj.items.all():
                        slug1 += "_" +item.slug
                obj.slug=slug1 
            except:
                obj.slug = obj.title
        super().save_model(request, obj, form, change)


 

def export_all_orders(modeladmin, request, queryset):
    ordersConfirmed = queryset.all()
    modeladmin.message_user(request, "Confirmed Orders Exported successfully.")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(['ORDER ID',"USERNAME","ITEMS ORDERED","CONFIRMED",'REF CODE', 'TOTAL AMOUNT','PAYEMENT METHODE','DEPOSIT AMOUNT','MESSAGE','IP ADDRESS','CARD NAME HOLDER','ORDERED DATE','PHONE N°','SHIPPING TYPE',"SHIPPING ADDRESS", 'WILATA SHIPPING','COMMUN SHIPPING','SHIPPING PRICE','CUPON'])
    item_orderd=[]
    for order in ordersConfirmed:
        item_orderd=[]
        for item in order.items.all():
            prod_cr=f"{item.item.title} * {item.quantity} with id {item.item.id} , size : {item.size}"
            item_orderd.append(prod_cr)
        item_orderd=str(item_orderd)
        ws.append([order.id,order.user.username,item_orderd,order.ordered, order.ref_code,order.total_amount,order.paiement_meth,order.depositAmount,order.message,order.ip_address,order.cardholderName,order.ordered_date,order.phone_number,order.shipping_type,order.shipping_address,order.wilaya_ship,order.commun_ship,order.shipping_price,order.coupon])

        # Prepare the response for Excel download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ALL_ORDERS.xlsx'
    wb.save(response)
    return response

export_all_orders.short_description = "Export All Orders"

def export_confirmed_orders(modeladmin, request, queryset):
    ordersConfirmed = queryset.filter(ordered=True)
    modeladmin.message_user(request, "Confirmed Orders Exported successfully.")
    wb = openpyxl.Workbook()
    ws = wb.active
    item_orderd=[]
     
    ws.append(['ORDER ID',"USERNAME","ITEMS ORDERED","CONFIRMED",'REF CODE', 'TOTAL AMOUNT','PAYEMENT METHODE','DEPOSIT AMOUNT','MESSAGE','IP ADDRESS','CARD NAME HOLDER','ORDERED DATE','PHONE N°','SHIPPING TYPE',"SHIPPING ADDRESS", 'WILATA SHIPPING','COMMUN SHIPPING','SHIPPING PRICE','CUPON'])

    for order in ordersConfirmed:
        item_orderd=[]
        for item in order.items.all():
            prod_cr=f"{item.item.title} * {item.quantity} with id {item.item.id} , size : {item.size}"
            item_orderd.append(prod_cr)
        item_orderd=str(item_orderd)
        ws.append([order.id,order.user.username, item_orderd,order.ordered,order.ref_code,order.total_amount,order.paiement_meth,order.depositAmount,order.message,order.ip_address,order.cardholderName,order.ordered_date,order.phone_number,order.shipping_type,order.shipping_address,order.wilaya_ship,order.commun_ship,order.shipping_price,order.coupon])

        # Prepare the response for Excel download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=CONFIRMED_ORDERS.xlsx'
    wb.save(response)
    return response

export_confirmed_orders.short_description = "Export Confirmed Orders"

class OrderConfirmAdmin(admin.ModelAdmin):
    list_display = ['user','ordered', 'ref_code','total_amount','received'] 
    list_filter = ['user', 'ref_code']
    search_fields = ['user', 'ref_code']
    actions = [export_confirmed_orders,export_all_orders]

 
from django.utils.safestring import mark_safe


class Imageitemdef(admin.ModelAdmin):
    list_display = ['get_image_display', 'item','item_category','item_id','image'] 
    list_filter = ['image', 'item', ]
    search_fields = ['item__title','item__id_item','item__category__title']
    actions = []

    def get_image_display(self, obj):
        return mark_safe(f'<img src="{obj.image.url}" width="50" height="50" />')

    get_image_display.allow_tags = True
    get_image_display.short_description = 'Image Preview'

    def item_id(self, obj):
        return obj.item.id_item 
    
    def item_category(self, obj):
        return obj.item.category 

class OrderItemfun(admin.ModelAdmin):
    list_display = [ 'user','ordered','Order_item_detail'] 
    def Order_item_detail(self, obj):
        return str(obj)
    
    
    list_filter = ['user']
    search_fields = ['user']

    
   
class slideAdmin(admin.ModelAdmin):
    list_display = ['get_image_display1','get_image_display2', 'caption1' ] 
    list_filter = ['caption1']
    search_fields = ['caption1']
   

    def get_image_display1(self, obj):
        return mark_safe(f'<img src="{obj.image.url}" width="80" height="60" />')
    def get_image_display2(self, obj):
        return mark_safe(f'<img src="{obj.image_mob.url}" width="50" height="60" />')

    get_image_display1.allow_tags = True
    get_image_display1.short_description = 'Image Web'
    get_image_display2.allow_tags = True
    get_image_display2.short_description = 'Image Mobile'
 

admin.site.register(GenderCategory,CategoryGenTopAdmin)
admin.site.register(ImageItem,Imageitemdef)
admin.site.register(Ad_homePage)
admin.site.register(Banner_category)
admin.site.register(Item, ItemAdmin)
admin.site.register(Category, CategoryAdmin)
admin.site.register(Slide,slideAdmin)
admin.site.register(Essential)
admin.site.register(OrderItem,OrderItemfun)
admin.site.register(Order,OrderConfirmAdmin)
admin.site.register(Coupon)
admin.site.register(WishList)
admin.site.register(TopCategory,CategoryGenTopAdmin)
# admin.site.register(ShopHeader)
admin.site.register(Matacor_info)
# admin.site.register(Images_upload)
admin.site.register(NewsLetterEmails)
admin.site.register(Comments_and_Ratings)
admin.site.register(Profile)
 
 